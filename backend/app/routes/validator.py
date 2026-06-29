from __future__ import annotations

import re
import tempfile
import unicodedata
import uuid
from io import BytesIO
from pathlib import Path
from typing import Annotated, Any

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from app.schemas.issues import ValidationIssue
from app.services.rule_engine import issue, validate_basic_row, validate_structure
from app.services.spreadsheet_reader import parse_xlsx
from app.services.geography_ibge_sqlite import reference_status as ibge_reference_status, validate_geography_ibge
from app.services.taxonomy_ffb_sqlite import reference_status as ffb_reference_status, validate_taxonomy_ffb

router = APIRouter()
LOCAL_JOBS: dict[str, dict[str, Any]] = {}

LOCAL_TAXA = {
    ("bertholletia", "excelsa"): {"family": "LECYTHIDACEAE", "accepted": True, "scientific_name": "Bertholletia excelsa"},
    ("poecilanthe", "parviflora"): {"family": "LEGUMINOSAE", "accepted": True, "scientific_name": "Poecilanthe parviflora"},
    ("libidibia", "ferrea"): {"family": "LEGUMINOSAE", "accepted": True, "scientific_name": "Libidibia ferrea"},
    ("centrolobium", "robustum"): {"family": "LEGUMINOSAE", "accepted": True, "scientific_name": "Centrolobium robustum"},
    ("chamaecrista", "ensiformis"): {"family": "LEGUMINOSAE", "accepted": False, "scientific_name": "Chamaecrista ensiformis", "accepted_name": "Chamaecrista ensiformis sensu verificar"},
}
LOCAL_GENERA = {genus for genus, _ in LOCAL_TAXA}
GENUS_SUGGESTIONS = {"chamaechrista": "Chamaecrista", "poecilanthes": "Poecilanthe", "libidibiaa": "Libidibia"}

STATE_BOXES = {
    "am": ( -10.5,  3.2, -74.2, -55.8),
    "amazonas": ( -10.5,  3.2, -74.2, -55.8),
}
MUNICIPALITY_BOXES = {
    "sao gabriel da cachoeira": (-1.8, 3.1, -70.6, -63.8),
    "são gabriel da cachoeira": (-1.8, 3.1, -70.6, -63.8),
}


def _norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text.strip().lower())


def _issue_to_dict(item: ValidationIssue) -> dict[str, Any]:
    return item.model_dump()


def _local_taxonomy(record: dict[str, Any]) -> list[ValidationIssue]:
    row = record.get("_row_number")
    family = record.get("family")
    genus = record.get("genus")
    sp1 = record.get("sp1")
    issues: list[ValidationIssue] = []

    if not genus:
        return issues

    genus_norm = _norm(genus)
    sp1_norm = _norm(sp1)

    if genus_norm not in LOCAL_GENERA:
        suggestion = GENUS_SUGGESTIONS.get(genus_norm)
        issues.append(issue(
            row, "genus", "error", "GENUS_NOT_FOUND_FFB",
            "Gênero não encontrado na base taxonômica local de teste.", genus,
            suggestion=suggestion,
            source="Flora e Funga do Brasil (fixture local de desenvolvimento)",
        ))
        return issues

    genus_records = [v for (g, _), v in LOCAL_TAXA.items() if g == genus_norm]
    expected_family = genus_records[0].get("family") if genus_records else None
    if family and expected_family and _norm(family) != _norm(expected_family):
        issues.append(issue(
            row, "family", "warning", "FAMILY_GENUS_MISMATCH",
            f"Família informada não coincide com a família esperada para o gênero no teste local ({expected_family}).",
            family, suggestion=expected_family,
            source="Flora e Funga do Brasil (fixture local de desenvolvimento)",
        ))

    if not sp1:
        return issues

    taxon = LOCAL_TAXA.get((genus_norm, sp1_norm))
    if not taxon:
        suggestions = [v["scientific_name"] for (g, _), v in LOCAL_TAXA.items() if g == genus_norm]
        issues.append(issue(
            row, "sp1", "error", "SPECIES_NOT_FOUND_FFB",
            "Espécie não encontrada na base taxonômica local de teste.", f"{genus} {sp1}",
            suggestion=", ".join(suggestions) if suggestions else None,
            source="Flora e Funga do Brasil (fixture local de desenvolvimento)",
        ))
        return issues

    if not taxon.get("accepted", True):
        issues.append(issue(
            row, "sp1", "warning", "TAXON_NOT_ACCEPTED",
            "Nome encontrado no teste local, mas marcado para revisão de aceitação taxonômica.",
            taxon.get("scientific_name"), suggestion=taxon.get("accepted_name"),
            source="Flora e Funga do Brasil (fixture local de desenvolvimento)",
        ))
    return issues


def _inside_box(lat: float, lon: float, box: tuple[float, float, float, float]) -> bool:
    min_lat, max_lat, min_lon, max_lon = box
    return min_lat <= lat <= max_lat and min_lon <= lon <= max_lon


def _local_geography(record: dict[str, Any]) -> list[ValidationIssue]:
    row = record.get("_row_number")
    lat = record.get("lat")
    lon = record.get("long")
    majorarea = record.get("majorarea")
    minorarea = record.get("minorarea")
    issues: list[ValidationIssue] = []

    if lat is None or lon is None:
        return issues

    state_key = _norm(majorarea)
    state_box = STATE_BOXES.get(state_key)
    if majorarea and not state_box:
        issues.append(issue(row, "majorarea", "warning", "STATE_NOT_FOUND_LOCAL", "Estado/área maior não está no fixture geográfico local de teste.", majorarea, source="PostGIS (fixture local de desenvolvimento)"))
    elif state_box and not _inside_box(float(lat), float(lon), state_box):
        issues.append(issue(row, "lat", "error", "POINT_OUTSIDE_STATE", "Coordenada não cai dentro do estado informado no teste geográfico local.", f"{lat}, {lon}", source="PostGIS (fixture local de desenvolvimento)"))

    muni_key = _norm(minorarea)
    muni_box = MUNICIPALITY_BOXES.get(muni_key)
    if minorarea and not muni_box:
        issues.append(issue(row, "minorarea", "warning", "MUNICIPALITY_NOT_FOUND_LOCAL", "Município/área menor não está no fixture geográfico local de teste.", minorarea, source="PostGIS (fixture local de desenvolvimento)"))
    elif muni_box and not _inside_box(float(lat), float(lon), muni_box):
        issues.append(issue(row, "long", "error", "POINT_OUTSIDE_MUNICIPALITY", "Coordenada não cai dentro do município informado no teste geográfico local.", f"{lat}, {lon}", source="PostGIS (fixture local de desenvolvimento)"))

    return issues


def _find_issue_column(issue_col: str | None, mapping: dict[str, str], raw_headers: list[str]) -> str | None:
    if not issue_col:
        return None
    if issue_col in raw_headers:
        return issue_col
    for raw, canonical in mapping.items():
        if canonical == issue_col:
            return raw
    return issue_col


def _annotated_workbook(job: dict[str, Any]) -> BytesIO:
    header_row = int(job["header_row"])
    mapping = job["mapping"]
    raw_headers = job["raw_headers"]

    try:
        wb = load_workbook(BytesIO(job["file_bytes"]))
        sheet_name = job.get("sheet_name")
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        using_original = True
    except Exception:
        # Fallback: create a clean workbook from parsed rows if the original workbook has
        # odd dimensions or unsupported worksheet structures. This keeps the download usable.
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Planilha anotada"
        using_original = False
        header_row = 1
        for col_idx, raw in enumerate(raw_headers, start=1):
            ws.cell(row=header_row, column=col_idx).value = raw
        for out_row, record in enumerate(job.get("table", []), start=2):
            raw = record.get("_raw") or {}
            for col_idx, raw_header in enumerate(raw_headers, start=1):
                ws.cell(row=out_row, column=col_idx).value = raw.get(raw_header)

    raw_to_col: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        raw = "" if cell.value is None else str(cell.value).strip()
        if raw:
            raw_to_col[raw] = col_idx

    fills = {
        "error": PatternFill("solid", fgColor="F8D7DA"),
        "warning": PatternFill("solid", fgColor="FFE5B4"),
        "info": PatternFill("solid", fgColor="E4ECD7"),
    }
    fonts = {
        "error": Font(color="8B1F3C", bold=True),
        "warning": Font(color="8B4A00", bold=True),
        "info": Font(color="2E3B24", bold=True),
    }
    icons = {"error": "✖", "warning": "⚠", "info": "ⓘ"}

    for item in job["issues"]:
        row = item.get("row_number")
        raw_col = _find_issue_column(item.get("column_name"), mapping, raw_headers)
        col = raw_to_col.get(raw_col) if raw_col else None
        if not row or not col:
            continue
        try:
            row_int = int(row)
        except Exception:
            continue
        if not using_original:
            # Original row numbers refer to the uploaded sheet. In fallback workbook, data
            # starts one line after the header and original header row may not be 1.
            row_int = row_int - int(job["header_row"]) + 1
            if row_int < 2:
                continue
        cell = ws.cell(row=row_int, column=int(col))
        sev = item.get("severity", "info")
        msg = f"{icons.get(sev, 'ⓘ')} {item.get('message', '')}"
        if item.get("suggestion"):
            msg += f" Sugestão: {item['suggestion']}"
        original = "" if cell.value is None else str(cell.value)
        cell.value = msg if not original.strip() else f"{original}\n{msg}"
        cell.fill = fills.get(sev, fills["info"])
        cell.font = fonts.get(sev, fonts["info"])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        comment_text = f"{item.get('code', '')}\n{item.get('message', '')}"
        if item.get("suggestion"):
            comment_text += f"\nSugestão: {item['suggestion']}"
        if item.get("source"):
            comment_text += f"\nFonte: {item['source']}"
        cell.comment = Comment(comment_text, "Tsiino Validator")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    if "Validação Tsiino" in wb.sheetnames:
        del wb["Validação Tsiino"]
    rep = wb.create_sheet("Validação Tsiino", 0)
    rep.append(["Resumo da validação"])
    rep.append(["Linhas", job["summary"].get("total_rows")])
    rep.append(["Erros", job["summary"].get("error_count")])
    rep.append(["Alertas", job["summary"].get("warning_count")])
    rep.append([])
    rep.append(["Linha", "Campo", "Tipo", "Código", "Mensagem", "Valor", "Sugestão", "Fonte"])
    for item in job["issues"]:
        rep.append([
            item.get("row_number"), item.get("column_name"), item.get("severity"), item.get("code"),
            item.get("message"), item.get("value"), item.get("suggestion"), item.get("source"),
        ])
    for row in rep.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    rep.column_dimensions["A"].width = 12
    rep.column_dimensions["B"].width = 18
    rep.column_dimensions["C"].width = 12
    rep.column_dimensions["D"].width = 28
    rep.column_dimensions["E"].width = 60
    rep.column_dimensions["F"].width = 25
    rep.column_dimensions["G"].width = 34
    rep.column_dimensions["H"].width = 34
    rep.freeze_panes = "A7"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@router.get("/status")
def status() -> dict[str, Any]:
    taxonomy = ffb_reference_status()
    return {
        "mode": "local-development",
        "taxonomy": taxonomy,
        "geography": ibge_reference_status(),
    }

@router.post("/upload")
async def upload_spreadsheet(
    file: Annotated[UploadFile, File(...)],
    validate_taxonomy: Annotated[bool, Form()] = True,
    validate_geography: Annotated[bool, Form()] = True,
    sheet_name: Annotated[str | None, Form()] = None,
) -> dict[str, str]:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Envie uma planilha .xlsx ou .xlsm.")

    job_id = str(uuid.uuid4())
    content = await file.read()

    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
        path = Path(tmpdir) / file.filename
        path.write_bytes(content)
        try:
            parsed = parse_xlsx(path, sheet_name=sheet_name or None)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Erro ao ler a planilha: {exc}") from exc

    all_issues: list[ValidationIssue] = []
    all_issues.extend(validate_structure(parsed.records, parsed.header.missing_minimum, parsed.header.unknown_headers))

    for record in parsed.records:
        all_issues.extend(validate_basic_row(record))
        if validate_taxonomy:
            all_issues.extend(validate_taxonomy_ffb(record))
        if validate_geography:
            all_issues.extend(validate_geography_ibge(record))

    if validate_taxonomy:
        all_issues.append(issue(None, None, "info", "TAXONOMY_LOCAL_FIXTURE", "Modo local: validação taxonômica demonstrativa com fixture pequeno. No deploy, usar Flora e Funga do Brasil importada em PostgreSQL.", source="Tsiino local mode"))
    if validate_geography:
        all_issues.append(issue(None, None, "info", "GEOGRAPHY_LOCAL_FIXTURE", "Modo local: validação geográfica demonstrativa com caixas aproximadas. No deploy, usar PostGIS com malhas oficiais.", source="Tsiino local mode"))

    issue_dicts = [_issue_to_dict(item) for item in all_issues]
    error_count = sum(1 for item in issue_dicts if item.get("severity") == "error")
    warning_count = sum(1 for item in issue_dicts if item.get("severity") == "warning")

    summary = {
        "id": job_id,
        "job_id": job_id,
        "status": "finished",
        "total_rows": len(parsed.records),
        "error_count": error_count,
        "warning_count": warning_count,
    }

    LOCAL_JOBS[job_id] = {
        "summary": summary,
        "issues": issue_dicts,
        "table": parsed.records,
        "file_bytes": content,
        "filename": file.filename,
        "sheet_name": sheet_name,
        "header_row": parsed.header.header_row,
        "mapping": parsed.header.mapping,
        "raw_headers": parsed.header.raw_headers,
    }

    return {"job_id": job_id, "status_url": f"/api/validator/jobs/{job_id}"}


@router.get("/jobs/{job_id}")
def get_job(job_id: str) -> dict[str, Any]:
    job = LOCAL_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado.")
    return job["summary"]


@router.get("/jobs/{job_id}/issues")
def get_issues(job_id: str) -> list[dict[str, Any]]:
    job = LOCAL_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado.")
    return job["issues"]


@router.get("/jobs/{job_id}/table")
def get_table(job_id: str) -> list[dict[str, Any]]:
    job = LOCAL_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado.")
    return job["table"]


@router.get("/jobs/{job_id}/download.xlsx")
def download_annotated(job_id: str) -> StreamingResponse:
    job = LOCAL_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado.")
    output = _annotated_workbook(job)
    filename = f"tsiino_planilha_anotada_{job_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/jobs/{job_id}/map.geojson")
def get_map(job_id: str) -> dict[str, Any]:
    job = LOCAL_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado.")

    error_rows = {item.get("row_number") for item in job["issues"] if item.get("severity") == "error"}
    features = []
    for record in job["table"]:
        lat = record.get("lat")
        lon = record.get("long")
        if lat is None or lon is None:
            continue
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [lon, lat]},
            "properties": {
                "row_number": record.get("_row_number"),
                "collector": record.get("collector"),
                "number": record.get("number"),
                "family": record.get("family"),
                "genus": record.get("genus"),
                "sp1": record.get("sp1"),
                "has_error": record.get("_row_number") in error_rows,
            },
        })
    return {"type": "FeatureCollection", "features": features}
