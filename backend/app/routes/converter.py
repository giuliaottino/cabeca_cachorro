from __future__ import annotations

import json
from typing import Optional

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import Response

from app.services.spreadsheet_converter import (
    build_template_workbook,
    convert_spreadsheet,
    preview_spreadsheet,
)

router = APIRouter(prefix='/converter', tags=['converter'])


@router.get('/template')
def download_template():
    data = build_template_workbook()
    return Response(
        data,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="template_inpa_brahms.xlsx"'},
    )


@router.post('/preview')
async def preview_converter(file: UploadFile = File(...), sheet_name: Optional[str] = Form(None)):
    try:
        data = await file.read()
        return preview_spreadsheet(data, sheet_name=sheet_name)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post('/convert')
async def convert_converter(
    file: UploadFile = File(...),
    mapping: Optional[str] = Form(None),
    mapping_json: Optional[str] = Form(None),
    sheet_name: Optional[str] = Form(None),
    include_recommendations: bool = Form(False),
):
    try:
        data = await file.read()
        raw_mapping = mapping_json or mapping or '{}'
        try:
            parsed_mapping = json.loads(raw_mapping) if raw_mapping else {}
        except Exception:
            parsed_mapping = {}
        out = convert_spreadsheet(
            data,
            mapping=parsed_mapping,
            sheet_name=sheet_name,
            include_recommendations=include_recommendations,
        )
        return Response(
            out,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="planilha_convertida_INPA.xlsx"'},
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

# TSIINO_PREVIEW_ROWS_V17
# Endpoint auxiliar para o mapeador: devolve as linhas reais da planilha enviada,
# sem depender da heuristica do preview principal.
import tempfile as _tempfile_v17
import os as _os_v17
from fastapi import UploadFile as _UploadFileV17, File as _FileV17, Form as _FormV17
from openpyxl import load_workbook as _load_workbook_v17


def _tsiino_v17_cell_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _tsiino_v17_header_score(values):
    texts = [_tsiino_v17_cell_text(v) for v in values]
    nonempty = [t for t in texts if t]
    if not nonempty:
        return -1
    long_penalty = sum(1 for t in nonempty if len(t) > 80)
    keyword_bonus = 0
    keys = " collector coletor number numero família familia genus genero country pais estado município municipio latitude longitude localidade "
    for t in nonempty:
        low = t.lower()
        if any(k in low for k in keys.split()):
            keyword_bonus += 1
    return len(nonempty) * 3 + keyword_bonus - long_penalty * 2


def _tsiino_v17_detect_header_row(ws, max_scan_rows=20):
    best_idx = 1
    best_score = -1
    max_row = min(ws.max_row or 1, max_scan_rows)
    max_col = ws.max_column or 1
    for idx in range(1, max_row + 1):
        values = [ws.cell(idx, col).value for col in range(1, max_col + 1)]
        score = _tsiino_v17_header_score(values)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


@router.post("/preview_rows")
async def preview_rows(file: UploadFile = File(...), sheet_name: str | None = Form(None), max_rows: int = Form(80)):
    """TSIINO_MAPPER_ROWS_HELP_V20

    Devolve colunas, linhas reais e recomendações do padrão INPA/BRAHMS para o mapeador.
    O contrato do frontend é deliberadamente redundante: source_columns é lista de strings,
    rows contém chaves brutas e chaves canônicas, e field_help traz as recomendações do template.
    """
    import tempfile
    import unicodedata
    import re as _re
    from pathlib import Path as _Path
    from openpyxl import load_workbook

    canonical_fields = [
        'accession', 'collector', 'prefix', 'number', 'suffix', 'addcoll',
        'colldd', 'collmm', 'collyy', 'initial', 'family', 'genus', 'detstatus',
        'sp1', 'rank1', 'sp2', 'detby', 'detdd', 'detmm', 'detyy', 'country',
        'majorarea', 'minorarea', 'gazetteer', 'locnotes', 'habitattxt', 'lat',
        'NS', 'long', 'EW', 'llunit', 'alt', 'alt1', 'plantdesc', 'vernacular',
        'dups', 'project', 'genbank'
    ]

    field_help = {
        'accession': 'Deixar este campo em branco. Aqui será incluído o número de registro no Herbário INPA.',
        'collector': 'Nome do coletor principal. Formato: Silva, LIL da. Inclua o nome completo na aba Pessoas-Coletores.',
        'prefix': 'Usar este campo somente no caso de séries diferentes. Não colocar suas iniciais.',
        'number': 'Número de coleta do coletor principal. Evite "s.n.". Use sistema simples com números sequenciais.',
        'suffix': 'Usar somente se o número de coleta foi repetido por engano. Nesse caso pode preencher com A, B, C...',
        'addcoll': 'Coletores adicionais. Demais pessoas presentes na coleta, no mesmo formato do coletor, separadas por ponto-vírgula.',
        'colldd': 'Dia da coleta.',
        'collmm': 'Mês da coleta.',
        'collyy': 'Ano da coleta.',
        'initial': 'Número de amostras por coleta, ou seja, em quantas amostras sua coleta foi dividida.',
        'family': 'Família. Verifique a grafia correta do nome da família na Flora e Funga do Brasil e/ou MOBOT.',
        'genus': 'Gênero. Verifique a grafia correta do nome do gênero na Flora e Funga do Brasil e/ou MOBOT.',
        'detstatus': 'Use “cf.” para conferir e “aff.” para material afim quando não tiver certeza do táxon.',
        'sp1': 'Epíteto específico. Se souber o gênero e não souber o epíteto, deixe em branco; não escreva “sp.”.',
        'rank1': 'Use “ssp.” para subespécie ou “var.” para variedade.',
        'sp2': 'Epíteto da subespécie ou variedade.',
        'detby': 'Nome do determinador, no mesmo formato do coletor. Separe múltiplos nomes por ponto-vírgula.',
        'detdd': 'Dia de determinação.',
        'detmm': 'Mês de determinação.',
        'detyy': 'Ano de determinação.',
        'country': 'País.',
        'majorarea': 'Estado por extenso. Não usar sigla.',
        'minorarea': 'Município.',
        'gazetteer': 'Localidade, por exemplo Campus do INPA, Reserva Florestal Adolfo Ducke, BR174 etc.',
        'locnotes': 'Detalhes de onde a coleta foi feita, por exemplo ao lado da cantina, atrás do alojamento etc.',
        'habitattxt': 'Descrição do tipo de habitat onde a planta foi coletada.',
        'lat': 'Latitude em graus decimais. Usar sinal negativo para Sul e sem sinal para Norte.',
        'NS': 'Utilizar N para Norte e S para Sul. Usar sinal negativo para Sul na latitude.',
        'long': 'Longitude em graus decimais. Usar sinal negativo para Oeste no Brasil.',
        'EW': 'Utilizar E para Leste e W para Oeste. Usar sinal negativo quando necessário.',
        'llunit': 'Deixar este campo em branco no padrão INPA/BRAHMS.',
        'alt': 'Altitude em metros. Não incluir “m”.',
        'alt1': 'Altitude máxima em metros. Use apenas quando houver duas medidas de altitude.',
        'plantdesc': 'Descrição detalhada do indivíduo coletado: hábito, tamanho, exsudatos, odores, cores etc. Não copiar descrições gerais.',
        'vernacular': 'Nome popular da planta, em minúsculo; se houver mais de um, separar por ponto-vírgula.',
        'dups': 'Siglas dos herbários onde o material será depositado. Para INPA, preencher com INPA.',
        'project': 'Campo opcional para nome do projeto e órgão financiador, com até 60 caracteres.',
        'genbank': 'Sequências, DNA ou gene. Se houver, mencionar gene, primers ou número de acesso no GenBank.',
    }

    aliases = {
        'accession': ['accession', 'registro inpa', 'tombo', 'numero registro', 'número registro', 'vai ser o numero de registro', 'número de registro no herbário inpa'],
        'collector': ['collector', 'coletor', 'coletores', 'coletor principal', 'nome do coletor principal', 'recordedby', 'recorded by'],
        'prefix': ['prefix', 'prefixo', 'serie', 'série', 'series diferentes'],
        'number': ['number', 'numero', 'número', 'numero de coleta', 'número de coleta', 'seu numero de coleta', 'seu número de coleta', 'record number', 'recordnumber'],
        'suffix': ['suffix', 'sufixo', 'numero de coleta foi repetido', 'número de coleta foi repetido'],
        'addcoll': ['addcoll', 'coletores adicionais', 'quem estava presente', 'demais pessoas', 'additional collectors'],
        'colldd': ['colldd', 'dia', 'dia da coleta'],
        'collmm': ['collmm', 'mes', 'mês', 'mes da coleta', 'mês da coleta'],
        'collyy': ['collyy', 'ano', 'ano da coleta'],
        'initial': ['initial', 'numero de amostras', 'número de amostras', 'n amostras'],
        'family': ['family', 'familia', 'família'],
        'genus': ['genus', 'genero', 'gênero'],
        'detstatus': ['detstatus', 'cf', 'aff'],
        'sp1': ['sp1', 'especie', 'espécie', 'epiteto da especie', 'epíteto da espécie', 'epiteto específico', 'epíteto específico'],
        'rank1': ['rank1', 'rank infraespecifico', 'subespecie', 'subespécie'],
        'sp2': ['sp2', 'variedade', 'epiteto infraespecifico', 'epíteto infraespecífico'],
        'detby': ['detby', 'determinador', 'nome do determinador'],
        'detdd': ['detdd', 'dia de determinacao', 'dia de determinação'],
        'detmm': ['detmm', 'mes de determinacao', 'mês de determinação'],
        'detyy': ['detyy', 'ano de determinacao', 'ano de determinação'],
        'country': ['country', 'pais', 'país'],
        'majorarea': ['majorarea', 'estado', 'estado por extenso', 'uf'],
        'minorarea': ['minorarea', 'municipio', 'município', 'cidade'],
        'gazetteer': ['gazetteer', 'localidade'],
        'locnotes': ['locnotes', 'detalhes de onde', 'notas localidade'],
        'habitattxt': ['habitattxt', 'habitat', 'ambiente', 'vegetacao', 'vegetação'],
        'lat': ['lat', 'latitude'],
        'NS': ['ns', 'n/s', 'hemisferio latitude', 'hemisfério latitude'],
        'long': ['long', 'longitude', 'lon'],
        'EW': ['ew', 'e/w', 'hemisferio longitude', 'hemisfério longitude'],
        'llunit': ['llunit', 'dms', 'graus minutos segundos', 'unidade coordenada'],
        'alt': ['alt', 'altitude'],
        'alt1': ['alt1', 'altitude maxima', 'altitude máxima'],
        'plantdesc': ['plantdesc', 'descricao da planta', 'descrição da planta', 'descricao detalhada', 'descrição detalhada'],
        'vernacular': ['vernacular', 'nome vernacular', 'nome popular'],
        'dups': ['dups', 'duplicatas', 'numero de duplicatas', 'número de duplicatas', 'duplicatas coletadas'],
        'project': ['project', 'projeto'],
        'genbank': ['genbank', 'gene', 'dna', 'sequencias', 'sequências'],
    }

    def norm(value):
        value = '' if value is None else str(value).strip()
        value = unicodedata.normalize('NFKD', value)
        value = ''.join(ch for ch in value if not unicodedata.combining(ch))
        value = value.lower()
        value = _re.sub(r'[^a-z0-9]+', ' ', value).strip()
        return value

    alias_to_field = {}
    for field, vals in aliases.items():
        alias_to_field[norm(field)] = field
        for v in vals:
            alias_to_field[norm(v)] = field
    for field in canonical_fields:
        alias_to_field[norm(field)] = field

    def guess_field(header):
        n = norm(header)
        if not n or n in {'exemplo', 'seus dados', 'dados'}:
            return None
        if n in alias_to_field:
            return alias_to_field[n]
        if 'notas para voce seguir' in n:
            return None
        # Containment fallback, useful for recommendation-like headers.
        best = None
        best_len = 0
        for alias, field in alias_to_field.items():
            if len(alias) < 4:
                continue
            if alias in n or n in alias:
                if len(alias) > best_len:
                    best = field
                    best_len = len(alias)
        return best

    data = await file.read()
    suffix = _Path(file.filename or 'uploaded.xlsx').suffix or '.xlsx'
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(data)
        tmp_path = tmp.name

    try:
        wb = load_workbook(tmp_path, data_only=True, read_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif 'Espécimes' in wb.sheetnames:
            ws = wb['Espécimes']
        elif 'Especimes' in wb.sheetnames:
            ws = wb['Especimes']
        else:
            ws = wb[wb.sheetnames[0]]

        max_col = min(ws.max_column or 1, 120)
        max_scan = min(ws.max_row or 1, 30)
        best_row = 1
        best_score = -10**9
        best_headers = []
        best_map = {}

        for row_idx in range(1, max_scan + 1):
            headers = [str(ws.cell(row_idx, c).value).strip() if ws.cell(row_idx, c).value is not None else '' for c in range(1, max_col + 1)]
            mapped = {}
            nonempty = [h for h in headers if h]
            if not nonempty:
                continue
            score = 0
            for ix, h in enumerate(headers):
                if not h:
                    continue
                field = guess_field(h)
                if field:
                    mapped[ix] = field
                    score += 6 if field in {'collector','number','family','genus','country','majorarea','minorarea'} else 3
                if len(h) > 120:
                    score -= 1
            score += min(len(nonempty), 40) * 0.1
            # Avoid choosing a pure data row over the recommendation/header row.
            data_like = sum(1 for h in nonempty if norm(h) in {'brasil', 'amazonas'} or _re.fullmatch(r'\d+(\.\d+)?', h))
            score -= data_like * 2
            if score > best_score:
                best_score = score
                best_row = row_idx
                best_headers = headers
                best_map = mapped

        # Unique source headers as strings; frontend relies on these exact keys.
        seen = {}
        source_columns = []
        source_columns_meta = []
        canonical_mapping = {}
        raw_to_canonical = {}
        for ix, header in enumerate(best_headers):
            if not header:
                continue
            base = header
            seen[base] = seen.get(base, 0) + 1
            name = base if seen[base] == 1 else f'{base} ({seen[base]})'
            field = best_map.get(ix)
            source_columns.append(name)
            source_columns_meta.append({
                'index': ix,
                'column_index': ix,
                'letter': ws.cell(best_row, ix + 1).column_letter,
                'key': name,
                'header': name,
                'name': name,
                'canonical': field,
            })
            if field and field not in canonical_mapping:
                canonical_mapping[field] = name
                raw_to_canonical[name] = field

        rows = []
        limit = max(1, min(int(max_rows or 80), 200))
        start = best_row + 1
        stop = min(ws.max_row or best_row, start + limit - 1)
        for r in range(start, stop + 1):
            raw_values = {}
            values_array = []
            canonical_values = {}
            has_any = False
            col_name_iter = iter(source_columns)
            for ix, header in enumerate(best_headers):
                if not header:
                    values_array.append('')
                    continue
                name = next(col_name_iter)
                val = ws.cell(r, ix + 1).value
                out = '' if val is None else str(val)
                if out != '':
                    has_any = True
                raw_values[name] = out
                values_array.append(out)
                field = best_map.get(ix)
                if field and field not in canonical_values:
                    canonical_values[field] = out
            if not has_any:
                continue
            row_obj = {'_row_number': r, '_source': raw_values, '_values': values_array}
            for f in canonical_fields:
                row_obj[f] = canonical_values.get(f, '')
            row_obj.update(raw_values)
            rows.append(row_obj)

        return {
            'status': 'ok',
            'sheet_name': ws.title,
            'header_row': best_row,
            'row_count': len(rows),
            'fields': [{'key': f, 'name': f.upper(), 'help': field_help.get(f, '')} for f in canonical_fields],
            'canonical_fields': canonical_fields,
            'field_help': field_help,
            'recommendations': field_help,
            'source_columns': source_columns,
            'source_columns_meta': source_columns_meta,
            'headers': source_columns,
            'mapping': raw_to_canonical,
            'canonical_mapping': canonical_mapping,
            'suggested_mapping': canonical_mapping,
            'rows': rows,
            'preview_rows': rows,
            'data_rows': rows,
        }
    finally:
        try:
            _Path(tmp_path).unlink(missing_ok=True)
        except Exception:
            pass

# TSIINO_CONVERT_TOLERANT_V24
# Endpoint alternativo tolerante para validar planilhas mapeadas.
# Ele nunca deve rejeitar a conversao por campo obrigatorio ausente: campos faltantes
# viram celulas em branco e a validacao normal aponta os erros depois.
try:
    from fastapi import File as _TsiinoFile, UploadFile as _TsiinoUploadFile, Request as _TsiinoRequest, HTTPException as _TsiinoHTTPException
    from fastapi.responses import StreamingResponse as _TsiinoStreamingResponse
    from openpyxl import load_workbook as _tsiino_load_workbook, Workbook as _TsiinoWorkbook
    from io import BytesIO as _TsiinoBytesIO
    import json as _tsiino_json
    import re as _tsiino_re
    import unicodedata as _tsiino_unicodedata
except Exception as _tsiino_import_error:  # pragma: no cover
    _tsiino_import_error_v24 = _tsiino_import_error

_TSIINO_CANONICAL_V24 = [
    'accession', 'collector', 'prefix', 'number', 'suffix', 'addcoll',
    'colldd', 'collmm', 'collyy', 'initial', 'family', 'genus', 'detstatus',
    'sp1', 'rank1', 'sp2', 'detby', 'detdd', 'detmm', 'detyy', 'country',
    'majorarea', 'minorarea', 'gazetteer', 'locnotes', 'habitattxt', 'lat',
    'NS', 'long', 'EW', 'llunit', 'alt', 'alt1', 'plantdesc', 'vernacular',
    'dups', 'project', 'genbank'
]

_TSIINO_REQUIRED_HINTS_V24 = {'collector', 'number', 'colldd', 'collmm', 'collyy', 'family', 'genus', 'country', 'majorarea', 'minorarea', 'lat', 'long', 'plantdesc'}

def _tsiino_norm_v24(value):
    if value is None:
        return ''
    text = str(value).strip()
    text = _tsiino_unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not _tsiino_unicodedata.combining(ch))
    text = text.lower()
    text = _tsiino_re.sub(r'[^a-z0-9]+', '', text)
    return text

_TSIINO_ALIAS_PATTERNS_V24 = [
    ('accession', ['accession', 'tombo', 'registroinpa', 'numeroderegistronoherbario', 'campobranco']),
    ('collector', ['collector', 'coletorprincipal', 'nomedocoletorprincipal', 'coletor', 'recordedby']),
    ('prefix', ['prefix', 'prefixo', 'serie', 'seriesdiferentes', 'naocolocaseusiniciais']),
    ('number', ['number', 'numero', 'numerodecoleta', 'seunumerodecoleta', 'recordnumber', 'collectionnumber', 'coletanumero']),
    ('suffix', ['suffix', 'sufixo', 'numerodecoletafoirepetido', 'repetidoporengano']),
    ('addcoll', ['addcoll', 'coletoresadicionais', 'quemestavapresentenacoleta', 'associatedcollectors', 'colaboradores']),
    ('colldd', ['colldd', 'diadacoleta', 'dia', 'day', 'hdiadia']),
    ('collmm', ['collmm', 'mesdacoleta', 'mes', 'month', 'lmes']),
    ('collyy', ['collyy', 'anodacoleta', 'ano', 'year', 'jano']),
    ('initial', ['initial', 'ndeamostras', 'numeroamostras', 'n', 'initials']),
    ('family', ['family', 'familia', 'familiaatento', 'familyname']),
    ('genus', ['genus', 'genero', 'mgenero', 'genericname']),
    ('detstatus', ['detstatus', 'cfaff', 'identificacao', 'determinacaostatus']),
    ('sp1', ['sp1', 'especie', 'epitetoespecifico', 'epiteto', 'species', 'scientificname']),
    ('rank1', ['rank1', 'subespecie', 'variedade', 'rankinfraespecifico']),
    ('sp2', ['sp2', 'epitetoinfraespecifico', 'infraespecifico']),
    ('detby', ['detby', 'determinador', 'nomedodeterminador', 'identifiedby']),
    ('detdd', ['detdd', 'diadeterminacao', 'diadadeterminacao']),
    ('detmm', ['detmm', 'mesdeterminacao', 'mesdadeterminacao']),
    ('detyy', ['detyy', 'anodeterminacao', 'anodadeterminacao']),
    ('country', ['country', 'pais', 'brasil']),
    ('majorarea', ['majorarea', 'estado', 'uf', 'estadoextenso', 'stateprovince']),
    ('minorarea', ['minorarea', 'municipio', 'municipioextenso', 'county', 'cidade']),
    ('gazetteer', ['gazetteer', 'localidade', 'localidadeextenso', 'specificlocality']),
    ('locnotes', ['locnotes', 'notaslocalidade', 'detalhesdeondeacoleta', 'localitynotes']),
    ('habitattxt', ['habitattxt', 'habitat', 'ambiente', 'vegetacao']),
    ('lat', ['lat', 'latitude']),
    ('NS', ['ns', 'norteousul', 'hemisferiolatitude']),
    ('long', ['long', 'longitude', 'lng']),
    ('EW', ['ew', 'lesteouoeste', 'hemisferiolongitude']),
    ('llunit', ['llunit', 'dms', 'grausminutossegundos', 'coordenadasem']),
    ('alt', ['alt', 'altitude', 'elevacao']),
    ('alt1', ['alt1', 'altitudemaxima']),
    ('plantdesc', ['plantdesc', 'descricaodaplanta', 'descricao', 'descricaoindividuo', 'observacao', 'epifita', 'arvoreta', 'subarbusto']),
    ('vernacular', ['vernacular', 'nomevernacular', 'nomepopular']),
    ('dups', ['dups', 'duplicatas', 'numerodeduplicatas', 'herbarios']),
    ('project', ['project', 'projeto']),
    ('genbank', ['genbank']),
]

_TSIINO_BAD_SOURCE_HINTS_V24 = ['notasparavoceseguir', 'useisso', 'consultelista', 'solteumacolunaaqui']

def _tsiino_guess_canonical_v24(header):
    n = _tsiino_norm_v24(header)
    if not n:
        return None
    for col in _TSIINO_CANONICAL_V24:
        if n == _tsiino_norm_v24(col):
            return col
    if any(bad in n for bad in _TSIINO_BAD_SOURCE_HINTS_V24):
        # algumas instrucoes ainda podem ser cabecalhos uteis; nao bloqueia se bater padrao especifico abaixo
        pass
    for canon, pats in _TSIINO_ALIAS_PATTERNS_V24:
        if any(p in n for p in pats):
            return canon
    return None

def _tsiino_cell_text_v24(v):
    if v is None:
        return ''
    if hasattr(v, 'isoformat'):
        try:
            return v.isoformat()
        except Exception:
            pass
    return str(v).strip()

def _tsiino_parse_mapping_v24(form):
    raw = None
    for key in ('mapping', 'mappings', 'column_mapping', 'columns'):
        if key in form:
            raw = form.get(key)
            break
    if raw is None:
        return {}
    if not isinstance(raw, str):
        raw = str(raw)
    try:
        obj = _tsiino_json.loads(raw)
    except Exception:
        return {}
    out = {}
    if isinstance(obj, dict):
        # formatos aceitos: {canonical: source}, {canonical: {source/header/name}}, {source: canonical}
        for k, v in obj.items():
            kk = str(k).strip()
            if isinstance(v, dict):
                vv = v.get('source') or v.get('header') or v.get('name') or v.get('label') or v.get('column') or v.get('value')
            else:
                vv = v
            if vv is None:
                continue
            vv = str(vv).strip()
            if kk in _TSIINO_CANONICAL_V24:
                out[kk] = vv
            else:
                guessed = _tsiino_guess_canonical_v24(vv) or _tsiino_guess_canonical_v24(kk)
                if guessed:
                    if _tsiino_guess_canonical_v24(vv) == guessed and kk not in _TSIINO_CANONICAL_V24:
                        out[guessed] = kk
                    else:
                        out[guessed] = vv
    elif isinstance(obj, list):
        for item in obj:
            if not isinstance(item, dict):
                continue
            canon = item.get('canonical') or item.get('target') or item.get('field') or item.get('inpa')
            source = item.get('source') or item.get('header') or item.get('name') or item.get('label') or item.get('column')
            if canon and source and str(canon) in _TSIINO_CANONICAL_V24:
                out[str(canon)] = str(source)
    return out

def _tsiino_header_score_v24(values):
    guessed = [_tsiino_guess_canonical_v24(v) for v in values if _tsiino_cell_text_v24(v)]
    unique = set(g for g in guessed if g)
    return len(unique) + len(unique & _TSIINO_REQUIRED_HINTS_V24) * 2

def _tsiino_detect_header_row_v24(ws):
    best_idx = 1
    best_score = -1
    max_row = min(ws.max_row or 1, 40)
    for idx in range(1, max_row + 1):
        values = [cell.value for cell in ws[idx]]
        score = _tsiino_header_score_v24(values)
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx

def _tsiino_build_source_tables_v24(ws, header_row):
    raw_headers = [_tsiino_cell_text_v24(cell.value) or f'COL_{i+1}' for i, cell in enumerate(ws[header_row])]
    # Garante unicidade de cabecalhos iguais/vazios
    seen = {}
    headers = []
    for h in raw_headers:
        base = h or 'COL'
        key = base
        if key in seen:
            seen[key] += 1
            key = f'{base}__{seen[base]}'
        else:
            seen[key] = 1
        headers.append(key)
    rows = []
    for ridx in range(header_row + 1, (ws.max_row or header_row) + 1):
        values = [_tsiino_cell_text_v24(cell.value) for cell in ws[ridx][:len(headers)]]
        if not any(values):
            continue
        rows.append({headers[i]: values[i] if i < len(values) else '' for i in range(len(headers))})
    return headers, rows

def _tsiino_auto_mapping_v24(headers):
    out = {}
    for h in headers:
        canon = _tsiino_guess_canonical_v24(h)
        if canon and canon not in out:
            out[canon] = h
    return out

def _tsiino_resolve_source_v24(source_name, headers):
    if not source_name:
        return None
    if source_name in headers:
        return source_name
    ns = _tsiino_norm_v24(source_name)
    if not ns:
        return None
    for h in headers:
        if _tsiino_norm_v24(h) == ns:
            return h
    # Labels truncados no frontend: usa inclusao normalizada curta com cuidado
    for h in headers:
        nh = _tsiino_norm_v24(h)
        if ns and (ns in nh or nh in ns):
            return h
    return None

@router.post('/convert_tolerant')
async def tsiino_convert_tolerant_v24(request: _TsiinoRequest, file: _TsiinoUploadFile = _TsiinoFile(...)):
    try:
        form = await request.form()
        content = await file.read()
        if not content:
            raise _TsiinoHTTPException(status_code=400, detail='Arquivo vazio.')
        wb = _tsiino_load_workbook(_TsiinoBytesIO(content), data_only=True)
        sheet_name = form.get('sheet') or form.get('sheet_name') or None
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row = _tsiino_detect_header_row_v24(ws)
        headers, source_rows = _tsiino_build_source_tables_v24(ws, header_row)
        submitted_mapping = _tsiino_parse_mapping_v24(form)
        auto_mapping = _tsiino_auto_mapping_v24(headers)
        final_mapping = dict(auto_mapping)
        # O mapeamento manual tem prioridade, mas so se conseguimos resolver a coluna fonte.
        for canon, src in submitted_mapping.items():
            resolved = _tsiino_resolve_source_v24(src, headers)
            if resolved:
                final_mapping[canon] = resolved
        out_wb = _TsiinoWorkbook()
        out_ws = out_wb.active
        out_ws.title = 'Espécimes'
        for cidx, canon in enumerate(_TSIINO_CANONICAL_V24, start=1):
            out_ws.cell(row=1, column=cidx, value=canon)
        for ridx, row in enumerate(source_rows, start=2):
            for cidx, canon in enumerate(_TSIINO_CANONICAL_V24, start=1):
                src = final_mapping.get(canon)
                value = row.get(src, '') if src else ''
                out_ws.cell(row=ridx, column=cidx, value=value)
        buf = _TsiinoBytesIO()
        out_wb.save(buf)
        buf.seek(0)
        out_name = (file.filename or 'planilha').rsplit('.', 1)[0] + '_convertida_INPA.xlsx'
        return _TsiinoStreamingResponse(
            buf,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{out_name}"'}
        )
    except _TsiinoHTTPException:
        raise
    except Exception as exc:
        raise _TsiinoHTTPException(status_code=400, detail=f'Não foi possível converter: {exc}')

# TSIINO_CONVERT_TOLERANT_V28
# Conversao tolerante e previsivel para o fluxo pos-mapeamento.
# Nao rejeita campo ausente; campos sem mapeamento ficam vazios e a validacao normal decide os erros.
try:
    from fastapi import File as _TsiinoFileV28, UploadFile as _TsiinoUploadFileV28, Request as _TsiinoRequestV28, HTTPException as _TsiinoHTTPExceptionV28
    from fastapi.responses import StreamingResponse as _TsiinoStreamingResponseV28
    from openpyxl import load_workbook as _tsiino_load_workbook_v28, Workbook as _TsiinoWorkbookV28
    from io import BytesIO as _TsiinoBytesIOV28
    import json as _tsiino_json_v28
    import re as _tsiino_re_v28
    import unicodedata as _tsiino_unicodedata_v28
except Exception as _tsiino_v28_import_error:  # pragma: no cover
    _tsiino_v28_import_error_saved = _tsiino_v28_import_error

_TSIINO_CANONICAL_V28 = [
    'accession', 'collector', 'prefix', 'number', 'suffix', 'addcoll',
    'colldd', 'collmm', 'collyy', 'initial', 'family', 'genus', 'detstatus',
    'sp1', 'rank1', 'sp2', 'detby', 'detdd', 'detmm', 'detyy', 'country',
    'majorarea', 'minorarea', 'gazetteer', 'locnotes', 'habitattxt', 'lat',
    'NS', 'long', 'EW', 'llunit', 'alt', 'alt1', 'plantdesc', 'vernacular',
    'dups', 'project', 'genbank'
]

_TSIINO_ALIASES_V28 = [
    ('accession', ['accession', 'registroinpa', 'tombo', 'numero de registro', 'deixa este campo em branco']),
    ('collector', ['collector', 'coletor principal', 'nome do coletor principal', 'coletor', 'recordedby']),
    ('prefix', ['prefix', 'prefixo', 'serie', 'series diferentes']),
    ('number', ['number', 'numero de coleta', 'número de coleta', 'seu numero de coleta', 'recordnumber']),
    ('suffix', ['suffix', 'sufixo', 'coleta foi repetido']),
    ('addcoll', ['addcoll', 'coletores adicionais', 'quem estava presente', 'associatedcollectors']),
    ('colldd', ['colldd', 'dia da coleta', 'dia']),
    ('collmm', ['collmm', 'mes da coleta', 'mês da coleta', 'mes', 'mês']),
    ('collyy', ['collyy', 'ano da coleta', 'ano']),
    ('initial', ['initial', 'n de amostras', 'nº de amostras', 'numero de amostras']),
    ('family', ['family', 'familia', 'família']),
    ('genus', ['genus', 'genero', 'gênero']),
    ('detstatus', ['detstatus', 'cfaff', 'cf/aff']),
    ('sp1', ['sp1', 'epiteto especifico', 'epíteto específico', 'especie', 'espécie', 'species']),
    ('rank1', ['rank1', 'rank infra', 'subespecie', 'variedade']),
    ('sp2', ['sp2', 'epiteto infra', 'infraespecifico']),
    ('detby', ['detby', 'determinador', 'nome do determinador']),
    ('detdd', ['detdd', 'dia da determinacao', 'dia de determinação']),
    ('detmm', ['detmm', 'mes da determinacao', 'mês da determinação']),
    ('detyy', ['detyy', 'ano da determinacao', 'ano da determinação']),
    ('country', ['country', 'pais', 'país']),
    ('majorarea', ['majorarea', 'estado', 'uf']),
    ('minorarea', ['minorarea', 'municipio', 'município']),
    ('gazetteer', ['gazetteer', 'localidade']),
    ('locnotes', ['locnotes', 'notas da localidade', 'detalhes de onde']),
    ('habitattxt', ['habitattxt', 'habitat', 'ambiente']),
    ('lat', ['lat', 'latitude']),
    ('NS', ['ns', 'n/s', 'norte sul']),
    ('long', ['long', 'longitude', 'lon']),
    ('EW', ['ew', 'e/w', 'leste oeste']),
    ('llunit', ['llunit', 'dms', 'graus minutos segundos']),
    ('alt', ['alt', 'altitude']),
    ('alt1', ['alt1', 'altitude maxima']),
    ('plantdesc', ['plantdesc', 'descricao da planta', 'descrição da planta']),
    ('vernacular', ['vernacular', 'nome vernacular', 'nome popular']),
    ('dups', ['dups', 'duplicatas', 'numero de duplicatas']),
    ('project', ['project', 'projeto']),
    ('genbank', ['genbank']),
]

_TSIINO_BAD_HEADER_HINTS_V28 = ['notasparavoceseguir', 'consultelista', 'useisso', 'solteumacolunaaqui']

def _tsiino_norm_v28(value):
    if value is None:
        return ''
    text = str(value).strip()
    text = _tsiino_unicodedata_v28.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not _tsiino_unicodedata_v28.combining(ch))
    text = text.lower()
    text = _tsiino_re_v28.sub(r'[^a-z0-9]+', '', text)
    return text

def _tsiino_text_v28(value):
    if value is None:
        return ''
    if hasattr(value, 'isoformat'):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value).strip()

def _tsiino_guess_v28(header):
    n = _tsiino_norm_v28(header)
    if not n:
        return None
    for c in _TSIINO_CANONICAL_V28:
        if n == _tsiino_norm_v28(c):
            return c
    # Nao usa celulas de recomendacao como origem, exceto se houver match tecnico direto acima.
    if any(x in n for x in _TSIINO_BAD_HEADER_HINTS_V28):
        return None
    for canon, words in _TSIINO_ALIASES_V28:
        for w in words:
            nw = _tsiino_norm_v28(w)
            if nw and (nw in n or n in nw):
                return canon
    return None

def _tsiino_header_score_v28(values):
    guessed = [_tsiino_guess_v28(v) for v in values if _tsiino_text_v28(v)]
    unique = {g for g in guessed if g}
    required = {'collector', 'number', 'family', 'genus', 'country', 'majorarea', 'minorarea'}
    return len(unique) + 2 * len(unique & required)

def _tsiino_detect_header_row_v28(ws):
    best_row, best_score = 1, -1
    for idx in range(1, min(ws.max_row or 1, 50) + 1):
        vals = [c.value for c in ws[idx]]
        score = _tsiino_header_score_v28(vals)
        if score > best_score:
            best_row, best_score = idx, score
    return best_row

def _tsiino_headers_rows_v28(ws, header_row):
    raw_headers = [_tsiino_text_v28(c.value) or f'COL_{i+1}' for i, c in enumerate(ws[header_row])]
    seen = {}
    headers = []
    for h in raw_headers:
        key = h or 'COL'
        if key in seen:
            seen[key] += 1
            key = f'{key}__{seen[key]}'
        else:
            seen[key] = 1
        headers.append(key)
    rows = []
    for ridx in range(header_row + 1, (ws.max_row or header_row) + 1):
        vals = [_tsiino_text_v28(c.value) for c in ws[ridx][:len(headers)]]
        if any(vals):
            rows.append({headers[i]: vals[i] if i < len(vals) else '' for i in range(len(headers))})
    return headers, rows

def _tsiino_parse_mapping_v28(form):
    raw = None
    for key in ('mapping', 'mapping_json', 'mappings', 'column_mapping', 'columns'):
        if key in form:
            raw = form.get(key)
            break
    if raw is None:
        return {}
    try:
        obj = _tsiino_json_v28.loads(str(raw))
    except Exception:
        return {}
    out = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            canon = str(k).strip()
            source = v
            if isinstance(v, dict):
                source = v.get('source') or v.get('header') or v.get('name') or v.get('label') or v.get('column') or v.get('value')
            if canon in _TSIINO_CANONICAL_V28 and source:
                out[canon] = str(source).strip()
            else:
                guessed = _tsiino_guess_v28(canon)
                if guessed and source:
                    out[guessed] = str(source).strip()
    elif isinstance(obj, list):
        for item in obj:
            if isinstance(item, dict):
                canon = item.get('canonical') or item.get('target') or item.get('field') or item.get('inpa')
                source = item.get('source') or item.get('header') or item.get('name') or item.get('label') or item.get('column')
                if canon in _TSIINO_CANONICAL_V28 and source:
                    out[str(canon)] = str(source).strip()
    return out

def _tsiino_resolve_source_v28(name, headers):
    if not name:
        return None
    if name in headers:
        return name
    n = _tsiino_norm_v28(name)
    for h in headers:
        if _tsiino_norm_v28(h) == n:
            return h
    for h in headers:
        nh = _tsiino_norm_v28(h)
        if n and len(n) >= 4 and (n in nh or nh in n):
            return h
    return None

def _tsiino_auto_mapping_v28(headers):
    out = {}
    for h in headers:
        canon = _tsiino_guess_v28(h)
        if canon and canon not in out:
            out[canon] = h
    return out

@router.post('/convert_tolerant_v28')
async def tsiino_convert_tolerant_v28(request: _TsiinoRequestV28, file: _TsiinoUploadFileV28 = _TsiinoFileV28(...)):
    content = await file.read()
    if not content:
        raise _TsiinoHTTPExceptionV28(status_code=400, detail='Arquivo vazio.')
    try:
        wb = _tsiino_load_workbook_v28(_TsiinoBytesIOV28(content), data_only=True)
    except Exception as exc:
        raise _TsiinoHTTPExceptionV28(status_code=400, detail=f'Nao foi possivel abrir a planilha: {exc}')
    form = await request.form()
    sheet_name = form.get('sheet_name') or form.get('sheet')
    if sheet_name and str(sheet_name) in wb.sheetnames:
        ws = wb[str(sheet_name)]
    else:
        ws = wb[wb.sheetnames[0]]
    header_row = _tsiino_detect_header_row_v28(ws)
    headers, rows = _tsiino_headers_rows_v28(ws, header_row)
    manual = _tsiino_parse_mapping_v28(form)
    auto = _tsiino_auto_mapping_v28(headers)
    mapping = {**auto, **manual}
    resolved = {canon: _tsiino_resolve_source_v28(src, headers) for canon, src in mapping.items() if canon in _TSIINO_CANONICAL_V28}

    out_wb = _TsiinoWorkbookV28()
    out_ws = out_wb.active
    out_ws.title = 'Espécimes'
    out_ws.append(_TSIINO_CANONICAL_V28)
    for row in rows:
        out_ws.append([row.get(resolved.get(canon), '') if resolved.get(canon) else '' for canon in _TSIINO_CANONICAL_V28])

    bio = _TsiinoBytesIOV28()
    out_wb.save(bio)
    bio.seek(0)
    filename = (getattr(file, 'filename', None) or 'planilha').rsplit('.', 1)[0] + '_convertida_INPA.xlsx'
    return _TsiinoStreamingResponseV28(
        bio,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )

# TSIINO_STANDARD_CHECK_V36
@router.post("/standard_check")
async def standard_check(file: UploadFile = File(...), sheet_name: str | None = Form(None)):
    """Retorna se a planilha enviada tem cabeçalho técnico INPA/BRAHMS.

    Este endpoint existe para impedir que planilhas padrão, como a Zavatin, sejam enviadas
    indevidamente ao mapeador só porque a primeira célula da linha de cabeçalho contém
    texto auxiliar como "Nome do campo no Brahms ->".
    """
    import tempfile
    from pathlib import Path as _Path
    from openpyxl import load_workbook
    from app.services.schema_detector import detect_header, is_strict_inpa_header_detection

    suffix = _Path(file.filename or 'uploaded.xlsx').suffix or '.xlsx'
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = _Path(tmp.name)
    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        detection = detect_header(ws)
        detected = bool(is_strict_inpa_header_detection(detection))
        return {
            'detected_standard': detected,
            'is_standard': detected,
            'header_row': getattr(detection, 'header_row', None),
            'mapped_fields': sorted(set(getattr(detection, 'mapping', {}).values())),
            'missing_minimum': getattr(detection, 'missing_minimum', []),
        }
    except Exception as exc:
        return {
            'detected_standard': False,
            'is_standard': False,
            'error': str(exc),
        }
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass

