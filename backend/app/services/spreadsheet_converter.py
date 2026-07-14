from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

CANONICAL_COLUMNS = [
    'accession', 'collector', 'prefix', 'number', 'suffix', 'addcoll',
    'colldd', 'collmm', 'collyy', 'initial', 'family', 'genus', 'detstatus',
    'sp1', 'rank1', 'sp2', 'detby', 'detdd', 'detmm', 'detyy', 'country',
    'majorarea', 'minorarea', 'gazetteer', 'locnotes', 'habitattxt', 'lat',
    'NS', 'long', 'EW', 'llunit', 'alt', 'alt1', 'plantdesc', 'vernacular',
    'dups', 'project', 'genbank'
]

REQUIRED_FIELDS = {
    'collector', 'number', 'colldd', 'collmm', 'collyy', 'family', 'genus',
    'country', 'majorarea', 'minorarea', 'lat', 'long', 'plantdesc'
}

FIELD_LABELS = {
    'accession': 'Registro INPA', 'collector': 'Coletor principal', 'prefix': 'Prefixo',
    'number': 'Número de coleta', 'suffix': 'Sufixo', 'addcoll': 'Coletores adicionais',
    'colldd': 'Dia da coleta', 'collmm': 'Mês da coleta', 'collyy': 'Ano da coleta',
    'initial': 'Nº de amostras', 'family': 'Família', 'genus': 'Gênero',
    'detstatus': 'cf./aff.', 'sp1': 'Epíteto específico', 'rank1': 'Rank infraespecífico',
    'sp2': 'Epíteto infraespecífico', 'detby': 'Determinador', 'detdd': 'Dia da determinação',
    'detmm': 'Mês da determinação', 'detyy': 'Ano da determinação', 'country': 'País',
    'majorarea': 'Estado', 'minorarea': 'Município', 'gazetteer': 'Localidade',
    'locnotes': 'Notas de localidade', 'habitattxt': 'Habitat', 'lat': 'Latitude',
    'NS': 'N/S', 'long': 'Longitude', 'EW': 'E/W', 'llunit': 'Unidade lat/long',
    'alt': 'Altitude', 'alt1': 'Altitude máxima', 'plantdesc': 'Descrição da planta',
    'vernacular': 'Nome vernacular', 'dups': 'Duplicatas', 'project': 'Projeto', 'genbank': 'GenBank'
}

FIELD_HELP = {
    'collector': 'Nome do coletor principal. Formato: Sobrenome, Iniciais.',
    'number': 'Número da coleta do coletor principal. Evite s.n.',
    'colldd': 'Dia da coleta, em número.', 'collmm': 'Mês da coleta, em número.',
    'collyy': 'Ano da coleta com quatro dígitos.', 'family': 'Família botânica.',
    'genus': 'Gênero botânico.', 'sp1': 'Epíteto específico, sem escrever sp.',
    'country': 'País da coleta.', 'majorarea': 'Estado por extenso.',
    'minorarea': 'Município.', 'gazetteer': 'Localidade principal.',
    'locnotes': 'Detalhes do ponto/local onde a coleta foi feita.',
    'habitattxt': 'Descrição do habitat.', 'lat': 'Latitude em graus decimais.',
    'long': 'Longitude em graus decimais.', 'plantdesc': 'Descrição detalhada da planta.',
}

RECOMMENDATIONS = [FIELD_LABELS.get(c, c) for c in CANONICAL_COLUMNS]

BAD_HEADER_HINTS = [
    'notas para voce seguir', 'recomendacoes no preenchimento', 'linha de exemplo',
]


def strip_accents(value: str) -> str:
    return ''.join(ch for ch in unicodedata.normalize('NFKD', value) if not unicodedata.combining(ch))


def norm(value: Any) -> str:
    if value is None:
        return ''
    s = str(value).strip()
    s = s.replace('\xa0', ' ')
    s = strip_accents(s).lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def compact(value: Any) -> str:
    return norm(value).replace(' ', '')


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.replace('\xa0', ' ').strip()
        return s if s else None
    return value


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def map_header_to_field(value: Any) -> str | None:
    t = norm(value)
    c = compact(value)
    if not t:
        return None
    if t in CANONICAL_COLUMNS:
        return t
    if c in CANONICAL_COLUMNS:
        return c

    # Não usar texto de orientação geral como coluna de dados.
    if any(h in t for h in BAD_HEADER_HINTS):
        return None

    # Regras específicas primeiro, para evitar falsos positivos.
    if 'graus minutos' in t or 'dms' in t or 'sistema de graus' in t:
        return 'llunit'
    if 'numero de duplicatas' in t or 'duplicatas coletadas' in t or t == 'duplicatas':
        return 'dups'
    if 'numero de coleta' in t or 'seu numero de coleta' in t or t == 'numero' or t == 'n coleta':
        return 'number'
    if 'nome do coletor principal' in t or t == 'coletor' or 'coletor principal' in t:
        return 'collector'
    if 'coletores adicionais' in t or 'demais pessoas' in t or t == 'addcoll':
        return 'addcoll'
    if t in {'dia', 'dia coleta'} or 'dia da coleta' in t:
        return 'colldd'
    if t in {'mes', 'mes coleta'} or 'mes da coleta' in t:
        return 'collmm'
    if t == 'ano' or 'ano da coleta' in t:
        return 'collyy'
    if 'numero de amostras' in t or 'amostras por coleta' in t:
        return 'initial'
    if t == 'familia' or t.startswith('familia ') or 'familia botanica' in t:
        return 'family'
    if t == 'genero' or t.startswith('genero ') or 'grafia correta do nome do genero' in t:
        return 'genus'
    if 'epiteto especifico' in t or 'epitito da especie' in t or 'epiteto da especie' in t:
        return 'sp1'
    if 'subespecie' in t or 'variedade' in t or 'infraespecific' in t:
        return 'sp2'
    if 'determinador' in t:
        return 'detby'
    if 'dia de determinacao' in t:
        return 'detdd'
    if 'mes de determinacao' in t:
        return 'detmm'
    if 'ano de determinacao' in t:
        return 'detyy'
    if t in {'pais', 'país'} or t == 'country':
        return 'country'
    if t.startswith('estado') or t == 'uf' or t == 'majorarea':
        return 'majorarea'
    if t.startswith('municipio') or t == 'minorarea':
        return 'minorarea'
    if t.startswith('localidade') or t == 'gazetteer':
        return 'gazetteer'
    if 'detalhes de onde' in t or 'notas de localidade' in t or t == 'locnotes':
        return 'locnotes'
    if 'tipo de habitat' in t or t == 'habitat' or t == 'habitattxt':
        return 'habitattxt'
    if 'latitude' in t:
        return 'lat'
    if 'longitude' in t:
        return 'long'
    if t in {'ns', 'n s', 'norte sul'}:
        return 'NS'
    if t in {'ew', 'e w', 'leste oeste'}:
        return 'EW'
    if t.startswith('altitude maxima') or 'maximo' in t and 'altitude' in t:
        return 'alt1'
    if t.startswith('altitude'):
        return 'alt'
    if 'descricao detalhada' in t or 'descricao da planta' in t or 'descrição detalhada' in str(value).lower():
        return 'plantdesc'
    if 'nome vernacular' in t or 'nome popular' in t:
        return 'vernacular'
    if 'herbario de deposito' in t or t == 'herbario' or t == 'herbarios':
        return 'dups'
    if 'projeto' in t or 'financiamento' in t:
        return 'project'
    if 'genbank' in t or 'sequencias' in t:
        return 'genbank'
    if 'prefixo' in t:
        return 'prefix'
    if 'sufixo' in t:
        return 'suffix'
    if 'cf' in t or 'aff' in t:
        return 'detstatus'
    if 'registro inpa' in t or 'numero de registro' in t:
        return 'accession'
    return None


def row_values(ws: Worksheet, row_idx: int) -> list[Any]:
    return [clean_cell(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]


def select_sheet(wb, sheet_name: str | None = None) -> Worksheet:
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    for preferred in ('Espécimes', 'Especimes', 'Specimens', 'Sheet1'):
        if preferred in wb.sheetnames:
            return wb[preferred]
    return wb[wb.sheetnames[0]]


@dataclass
class HeaderInfo:
    row: int
    exact_count: int
    mapped_count: int
    mapping: dict[str, int]  # field -> 1-based source column
    headers: list[Any]


def score_header_row(ws: Worksheet, row_idx: int) -> HeaderInfo:
    headers = row_values(ws, row_idx)
    mapping: dict[str, int] = {}
    exact_count = 0
    mapped_count = 0
    for i, raw in enumerate(headers, start=1):
        if is_blank(raw):
            continue
        n = norm(raw)
        field = map_header_to_field(raw)
        if field:
            mapped_count += 1
            # Se o campo já existe, fica com o cabeçalho mais curto/específico.
            prev = mapping.get(field)
            if prev is None or len(str(raw)) < len(str(headers[prev - 1] or '')):
                mapping[field] = i
            if n in CANONICAL_COLUMNS:
                exact_count += 1
    return HeaderInfo(row_idx, exact_count, len(set(mapping)), mapping, headers)


def detect_header(ws: Worksheet) -> HeaderInfo:
    candidates = [score_header_row(ws, r) for r in range(1, min(ws.max_row, 15) + 1)]
    def key(info: HeaderInfo):
        required_hits = len(set(info.mapping) & REQUIRED_FIELDS)
        avg_len = sum(len(str(v or '')) for v in info.headers) / max(1, len([v for v in info.headers if not is_blank(v)]))
        long_penalty = 1 if avg_len > 80 else 0
        return (info.exact_count >= 8, info.exact_count * 4 + required_hits * 2 + info.mapped_count - long_penalty, -info.row)
    best = max(candidates, key=key)
    if best.mapped_count < 3:
        raise ValueError('Não foi possível detectar uma linha de cabeçalho na planilha enviada.')
    return postprocess_mapping(ws, best)


def col_samples(ws: Worksheet, col_idx: int, start_row: int, limit: int = 8) -> list[Any]:
    vals = []
    for r in range(start_row, min(ws.max_row, start_row + limit - 1) + 1):
        v = clean_cell(ws.cell(r, col_idx).value)
        if not is_blank(v):
            vals.append(v)
    return vals


def as_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(',', '.').strip()
    m = re.search(r'-?\d+(?:\.\d+)?', s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def infer_numeric_col(ws: Worksheet, start_row: int, used_cols: set[int], kind: str) -> int | None:
    best = None
    best_score = -1
    for col in range(1, ws.max_column + 1):
        if col in used_cols:
            continue
        nums = [as_float(ws.cell(r, col).value) for r in range(start_row, min(ws.max_row, start_row + 20) + 1)]
        nums = [x for x in nums if x is not None]
        if len(nums) < 3:
            continue
        if kind == 'lat':
            score = sum(1 for x in nums if -35 <= x <= 10)
        else:
            score = sum(1 for x in nums if -80 <= x <= -30 or 30 <= x <= 80)
        if score > best_score:
            best_score = score
            best = col
    return best if best_score >= 3 else None


def postprocess_mapping(ws: Worksheet, info: HeaderInfo) -> HeaderInfo:
    mapping = dict(info.mapping)
    headers = info.headers
    def h(field):
        c = mapping.get(field)
        return norm(headers[c - 1]) if c else ''
    # Corrige falsos positivos comuns nas planilhas de campo.
    if 'duplicatas' in h('collyy') or 'numero de duplicatas' in h('collyy'):
        mapping['dups'] = mapping.pop('collyy')
    if h('initial') == 'ano' or 'ano da coleta' in h('initial'):
        mapping['collyy'] = mapping.pop('initial')
    if 'epiteto' in h('alt'):
        mapping['sp1'] = mapping.pop('alt')
    if 'notas para voce seguir' in h('plantdesc'):
        mapping.pop('plantdesc', None)
    if 'sistema de graus' in h('long') or 'dms' in h('long'):
        mapping['llunit'] = mapping.pop('long')
    used = set(mapping.values())
    start_row = info.row + 1
    if 'lat' not in mapping:
        col = infer_numeric_col(ws, start_row, used, 'lat')
        if col:
            mapping['lat'] = col
            used.add(col)
    if 'long' not in mapping:
        col = infer_numeric_col(ws, start_row, used, 'long')
        if col:
            mapping['long'] = col
            used.add(col)
    return HeaderInfo(info.row, info.exact_count, len(set(mapping)), mapping, headers)


def is_standard(info: HeaderInfo) -> bool:
    # Padrão INPA/BRAHMS só quando a própria linha de cabeçalho é técnica.
    # Descrições/aliases como "Nome do coletor principal" ou "Estado por extenso"
    # NÃO devem ser tratadas como padrão; elas entram no mapeador.
    fields = set(info.mapping)
    ordered_hits = 0
    for raw, canonical in zip(info.headers, CANONICAL_COLUMNS):
        if norm(raw) == canonical:
            ordered_hits += 1
    required_core = {'accession', 'collector', 'number', 'family', 'genus', 'country', 'majorarea', 'minorarea', 'lat', 'long'}
    return info.exact_count >= 25 and ordered_hits >= 20 and required_core.issubset(fields)


def parse_mapping_json(value: Any) -> dict[str, Any]:
    if not value:
        return {}
    if isinstance(value, dict):
        return value
    try:
        return json.loads(str(value))
    except Exception:
        return {}


def mapping_to_indices(mapping: dict[str, Any], fallback: HeaderInfo) -> dict[str, int]:
    out: dict[str, int] = {}
    src = mapping or {}
    if 'mapping' in src and isinstance(src['mapping'], dict):
        src = src['mapping']
    for field, val in src.items():
        if field not in CANONICAL_COLUMNS or val in (None, ''):
            continue
        if isinstance(val, int):
            out[field] = val
        else:
            sval = str(val).strip()
            if re.fullmatch(r'[A-Za-z]+', sval):
                out[field] = column_index_from_string(sval.upper())
            elif sval.isdigit():
                out[field] = int(sval)
    # Preenche com sugestão se o usuário não mexeu naquele campo.
    for field, idx in fallback.mapping.items():
        out.setdefault(field, idx)
    return out


def preview_spreadsheet(file_bytes: bytes, sheet_name: str | None = None) -> dict[str, Any]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = select_sheet(wb, sheet_name)
    info = detect_header(ws)
    data_start = info.row + 1
    source_columns = []
    for col in range(1, ws.max_column + 1):
        raw = clean_cell(ws.cell(info.row, col).value)
        if is_blank(raw):
            continue
        letter = get_column_letter(col)
        mapped = None
        for field, idx in info.mapping.items():
            if idx == col:
                mapped = field
                break
        source_columns.append({
            'letter': letter,
            'index': col,
            'header': raw,
            'mapped_field': mapped,
            'samples': col_samples(ws, col, data_start, 5),
        })
    rows = []
    for r in range(data_start, min(ws.max_row, data_start + 19) + 1):
        rows.append({
            'row_number': r,
            'values': [clean_cell(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)],
        })
    fields = [{
        'name': c,
        'key': c,
        'label': FIELD_LABELS.get(c, c),
        'required': c in REQUIRED_FIELDS,
        'help': FIELD_HELP.get(c, FIELD_LABELS.get(c, c)),
    } for c in CANONICAL_COLUMNS]
    mapping_letters = {field: get_column_letter(idx) for field, idx in info.mapping.items()}
    return {
        'detected_standard': is_standard(info),
        'sheet_name': ws.title,
        'header_row': info.row,
        'data_start_row': data_start,
        'fields': fields,
        'standard_fields': fields,
        'source_columns': source_columns,
        'columns': source_columns,
        'rows': rows,
        'preview_rows': rows,
        'mapping': mapping_letters,
        'suggested_mapping': mapping_letters,
        'field_to_column': mapping_letters,
        'missing_required': [c for c in REQUIRED_FIELDS if c not in info.mapping],
    }


def row_is_empty_for_mapping(ws: Worksheet, r: int, mapping: dict[str, int]) -> bool:
    meaningful = ['collector', 'number', 'family', 'genus', 'country', 'majorarea', 'minorarea', 'lat', 'long', 'plantdesc']
    vals = [ws.cell(r, mapping[f]).value for f in meaningful if f in mapping]
    return all(is_blank(v) for v in vals)


def convert_spreadsheet(file_bytes: bytes, mapping: dict[str, Any] | None = None, sheet_name: str | None = None, *, include_recommendations: bool = False) -> bytes:
    wb_in = load_workbook(BytesIO(file_bytes), data_only=True)
    ws_in = select_sheet(wb_in, sheet_name)
    info = detect_header(ws_in)
    idx_map = mapping_to_indices(mapping or {}, info)

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = 'Espécimes'
    if include_recommendations:
        ws.append(RECOMMENDATIONS)
    ws.append(CANONICAL_COLUMNS)
    data_start = info.row + 1
    written = 0
    for r in range(data_start, ws_in.max_row + 1):
        if row_is_empty_for_mapping(ws_in, r, idx_map):
            continue
        row = []
        for field in CANONICAL_COLUMNS:
            col = idx_map.get(field)
            row.append(clean_cell(ws_in.cell(r, col).value) if col else None)
        # Ignora linha de exemplo se existir.
        joined = norm(' '.join(str(x or '') for x in row[:6]))
        if 'linha de exemplo' in joined:
            continue
        ws.append(row)
        written += 1
    # Aba de auditoria do mapeamento.
    audit = wb_out.create_sheet('Mapeamento')
    audit.append(['Campo INPA/BRAHMS', 'Coluna de origem', 'Cabeçalho de origem'])
    for field in CANONICAL_COLUMNS:
        col = idx_map.get(field)
        audit.append([field, get_column_letter(col) if col else None, clean_cell(ws_in.cell(info.row, col).value) if col else None])
    for col in range(1, len(CANONICAL_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    bio = BytesIO()
    wb_out.save(bio)
    return bio.getvalue()


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Espécimes'
    ws.append(RECOMMENDATIONS)
    ws.append(CANONICAL_COLUMNS)
    for col in range(1, len(CANONICAL_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
