from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.normalization import clean_cell, normalize_key, to_float
from app.services.schema_detector import CANONICAL_COLUMNS, COLUMN_ALIASES, REQUIRED_MINIMUM, HeaderDetection


class ParsedSpreadsheet:
    def __init__(self, header: HeaderDetection, records: list[dict[str, Any]]):
        self.header = header
        self.records = records


def _map_header(value: Any) -> str | None:
    raw = clean_cell(value)
    if raw is None:
        return None
    if raw in CANONICAL_COLUMNS:
        return raw
    key = normalize_key(raw)
    if key in CANONICAL_COLUMNS:
        return key
    return COLUMN_ALIASES.get(key)


def _safe_cell(value: Any) -> Any:
    return clean_cell(value)


def _row_values(row: tuple[Any, ...]) -> list[Any]:
    return [_safe_cell(value) for value in row]


def _detect_header_from_rows(rows: list[list[Any]], max_scan_rows: int = 20) -> HeaderDetection:
    best: HeaderDetection | None = None
    best_score = -1

    scan_limit = min(len(rows), max_scan_rows)
    for row_idx0 in range(scan_limit):
        values = rows[row_idx0]
        raw_headers = [str(v).strip() if v is not None else '' for v in values]
        mapped: dict[str, str] = {}
        unknown: list[str] = []

        for raw in raw_headers:
            if not raw:
                continue
            canonical = _map_header(raw)
            if canonical:
                mapped[raw] = canonical
            else:
                # Text-heavy instruction rows in INPA templates should not dominate.
                if len(raw) <= 80:
                    unknown.append(raw)

        mapped_values = set(mapped.values())
        score = len(mapped_values & set(REQUIRED_MINIMUM)) * 3 + len(mapped_values & set(CANONICAL_COLUMNS))
        if score > best_score:
            missing = [col for col in REQUIRED_MINIMUM if col not in mapped_values]
            best = HeaderDetection(
                header_row=row_idx0 + 1,
                mapping=mapped,
                raw_headers=raw_headers,
                missing_minimum=missing,
                unknown_headers=unknown,
            )
            best_score = score

    if best is None or best_score < 3:
        raise ValueError('Não foi possível detectar a linha de cabeçalho da planilha.')
    return best


def _looks_like_template_example(row_idx: int, header_row: int, raw: dict[str, Any]) -> bool:
    if row_idx != header_row + 1:
        return False
    joined = ' '.join(str(v).strip().lower() for v in raw.values() if v is not None)
    example_markers = ['a. coletor', 'b. auxiliar', 'exemplo', 'encholirium horridum', 'bromeliaceae encholirium']
    return any(marker in joined for marker in example_markers)


def _open_workbook(path: Path):
    # read_only=True avoids a class of openpyxl errors in files with odd worksheet dimensions.
    # If a workbook cannot be streamed, fall back to normal mode.
    try:
        return load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return load_workbook(path, data_only=True, read_only=False)


def _select_sheet(wb, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            available = ', '.join(wb.sheetnames)
            raise ValueError(f"A aba '{sheet_name}' não foi encontrada. Abas disponíveis: {available}")
        return wb[sheet_name]
    # Prefer the INPA/BRAHMS sheet when present, otherwise use the first sheet.
    for preferred in ('Espécimes', 'Especimes', 'Specimens', 'Sheet1'):
        if preferred in wb.sheetnames:
            return wb[preferred]
    return wb[wb.sheetnames[0]]


def parse_xlsx(path: Path, sheet_name: str | None = None) -> ParsedSpreadsheet:
    wb = _open_workbook(path)
    try:
        ws = _select_sheet(wb, sheet_name)

        rows: list[list[Any]] = []
        # Avoid ws.max_row/ws[1] because some browser-edited or legacy spreadsheets expose None-like dimensions.
        # IMPORTANT on Windows: read all rows before returning and close the workbook in finally,
        # otherwise TemporaryDirectory cleanup can fail with WinError 32 because openpyxl keeps
        # the uploaded .xlsx file handle open.
        for row in ws.iter_rows(values_only=True):
            values = _row_values(row)
            # Keep instruction rows; header detection needs them for row indices, but trim trailing empty cells.
            while values and values[-1] is None:
                values.pop()
            rows.append(values)
            if len(rows) >= 5000:
                break
    finally:
        close = getattr(wb, 'close', None)
        if callable(close):
            close()

    if not rows:
        raise ValueError('A planilha está vazia.')

    header = _detect_header_from_rows(rows)
    header_idx0 = header.header_row - 1
    header_values = header.raw_headers
    raw_to_canonical = header.mapping
    useful_header_count = max(1, len([h for h in header_values if h]))

    records: list[dict[str, Any]] = []
    for row_idx0 in range(header_idx0 + 1, len(rows)):
        row_idx = row_idx0 + 1
        values = rows[row_idx0]
        raw: dict[str, Any] = {}
        canonical: dict[str, Any] = {'_row_number': row_idx, '_raw': raw}
        empty_count = 0

        for pos, raw_header in enumerate(header_values):
            if not raw_header:
                continue
            value = values[pos] if pos < len(values) else None
            value = _safe_cell(value)
            raw[raw_header] = value
            if value is None:
                empty_count += 1
            canonical_name = raw_to_canonical.get(raw_header)
            if canonical_name:
                canonical[canonical_name] = value

        if empty_count >= useful_header_count - 1:
            continue
        if _looks_like_template_example(row_idx, header.header_row, raw):
            continue

        canonical['lat'] = to_float(canonical.get('lat'))
        canonical['long'] = to_float(canonical.get('long'))
        records.append(canonical)

    return ParsedSpreadsheet(header=header, records=records)
