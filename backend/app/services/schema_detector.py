from dataclasses import dataclass
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.services.normalization import clean_cell, normalize_key

# Campos canônicos da planilha INPA/BRAHMS observada no template.
CANONICAL_COLUMNS = [
    'accession', 'collector', 'prefix', 'number', 'suffix', 'addcoll',
    'colldd', 'collmm', 'collyy', 'initial', 'family', 'genus', 'detstatus',
    'sp1', 'rank1', 'sp2', 'detby', 'detdd', 'detmm', 'detyy', 'country',
    'majorarea', 'minorarea', 'gazetteer', 'locnotes', 'habitattxt', 'lat',
    'NS', 'long', 'EW', 'llunit', 'alt', 'alt1', 'plantdesc', 'vernacular',
    'dups', 'project', 'genbank'
]

# Aliases tolerados para aceitar pequenas modificações sem perder o significado.
COLUMN_ALIASES = {
    'numtombo': 'accession',
    'tombo': 'accession',
    'coletor': 'collector',
    'coletores': 'collector',
    'numero': 'number',
    'numcoleta': 'number',
    'numero_coleta': 'number',
    'sufixo': 'suffix',
    'coletores_adicionais': 'addcoll',
    'add_collector': 'addcoll',
    'dia': 'colldd',
    'dia_coleta': 'colldd',
    'mes': 'collmm',
    'mes_coleta': 'collmm',
    'ano': 'collyy',
    'ano_coleta': 'collyy',
    'familia': 'family',
    'genero': 'genus',
    'especie': 'sp1',
    'epiteto': 'sp1',
    'epiteto_especifico': 'sp1',
    'autor': 'author1',
    'author1': 'author1',
    'pais': 'country',
    'estado': 'majorarea',
    'uf': 'majorarea',
    'municipio': 'minorarea',
    'localidade': 'gazetteer',
    'notas_localidade': 'locnotes',
    'habitat': 'habitattxt',
    'latitude': 'lat',
    'longitude': 'long',
    'longitud': 'long',
    'unidade_latlong': 'llunit',
    'altitude': 'alt',
    'descricao': 'plantdesc',
    'descricao_planta': 'plantdesc',
    'nome_popular': 'vernacular',
    'duplicatas': 'dups',
    'herbario': 'dups',
    'herbarios': 'dups',
    'projeto': 'project',
}

REQUIRED_MINIMUM = [
    'collector', 'number', 'colldd', 'collmm', 'collyy', 'family', 'genus',
    'country', 'majorarea', 'minorarea', 'lat', 'long', 'plantdesc'
]


@dataclass
class HeaderDetection:
    header_row: int
    mapping: dict[str, str]
    raw_headers: list[str]
    missing_minimum: list[str]
    unknown_headers: list[str]


def _map_header(value: Any) -> str | None:
    raw = clean_cell(value)
    if raw is None:
        return None
    if raw in CANONICAL_COLUMNS:
        return raw
    key = normalize_key(raw)
    if key in CANONICAL_COLUMNS:
        return key
    return COLUMN_ALIASES.get(key)


def detect_header(ws: Worksheet, max_scan_rows: int = 12) -> HeaderDetection:
    best: HeaderDetection | None = None
    best_score = -1

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        raw_headers = [clean_cell(cell.value) or '' for cell in ws[row_idx]]
        mapped: dict[str, str] = {}
        unknown: list[str] = []

        for raw in raw_headers:
            if not raw:
                continue
            canonical = _map_header(raw)
            if canonical:
                mapped[raw] = canonical
            else:
                unknown.append(raw)

        score = len(set(mapped.values()) & set(REQUIRED_MINIMUM)) + len(set(mapped.values()) & set(CANONICAL_COLUMNS))
        if score > best_score:
            missing = [col for col in REQUIRED_MINIMUM if col not in set(mapped.values())]
            best = HeaderDetection(row_idx, mapped, raw_headers, missing, unknown)
            best_score = score

    if best is None or best_score < 3:
        raise ValueError('Não foi possível detectar a linha de cabeçalho da planilha.')
    return best

# TSIINO_SCHEMA_DETECTOR_V34
# Detector conservador de planilha padrão INPA/BRAHMS.
# Regra: só considera padrão quando encontra uma linha com campos técnicos canônicos exatos
# (accession, collector, number, family, genus, ...). Cabeçalhos descritivos continuam indo
# para o mapeador pré-validação.
def _tsiino_norm_header_v34(value):
    try:
        import unicodedata, re
        if value is None:
            return ''
        s = str(value).strip().replace('\ufeff', '')
        s = unicodedata.normalize('NFKD', s)
        s = ''.join(ch for ch in s if not unicodedata.combining(ch))
        s = s.lower()
        s = re.sub(r'[^a-z0-9]+', '_', s).strip('_')
        return s
    except Exception:
        return str(value or '').strip().lower()

_CANONICAL_SET_V34 = set(CANONICAL_COLUMNS)
_CANONICAL_NORM_TO_CANONICAL_V34 = {_tsiino_norm_header_v34(c): c for c in CANONICAL_COLUMNS}
_EXTRA_ALIASES_V34 = {
    'familia': 'family', 'genero': 'genus', 'especie': 'sp1', 'epiteto_especifico': 'sp1',
    'pais': 'country', 'estado': 'majorarea', 'municipio': 'minorarea',
    'latitude': 'lat', 'longitude': 'long', 'descricao_da_planta': 'plantdesc',
}
try:
    COLUMN_ALIASES.update(_EXTRA_ALIASES_V34)
except Exception:
    pass

def _tsiino_map_header_v34(raw, allow_alias=True):
    if raw is None:
        return None
    s = str(raw).strip().replace('\ufeff', '')
    if not s:
        return None
    if s in _CANONICAL_SET_V34:
        return s
    key = _tsiino_norm_header_v34(s)
    if key in _CANONICAL_NORM_TO_CANONICAL_V34:
        return _CANONICAL_NORM_TO_CANONICAL_V34[key]
    if allow_alias:
        return COLUMN_ALIASES.get(key)
    return None

def _tsiino_detect_header_from_values_v34(rows, max_scan_rows=20):
    best = None
    best_score = -1
    best_exact = -1
    required = set(REQUIRED_MINIMUM)
    canonical_set = set(CANONICAL_COLUMNS)
    for row_idx, row in enumerate(rows[:max_scan_rows], start=1):
        raw_headers = []
        mapped = {}
        unknown = []
        exact_count = 0
        for cell in row:
            raw = '' if cell is None else str(cell).strip().replace('\ufeff', '')
            raw_headers.append(raw)
            if not raw:
                continue
            canonical_exact = _tsiino_map_header_v34(raw, allow_alias=False)
            canonical_alias = canonical_exact or _tsiino_map_header_v34(raw, allow_alias=True)
            if canonical_exact:
                exact_count += 1
            if canonical_alias:
                mapped.setdefault(raw, canonical_alias)
            else:
                if _tsiino_norm_header_v34(raw) not in {'nome_do_campo_no_brahms', 'nome_do_campo_no_brahms_'}:
                    unknown.append(raw)
        mapped_values = set(mapped.values())
        score = len(mapped_values & required) * 5 + len(mapped_values & canonical_set) + exact_count * 3
        if score > best_score:
            missing = [col for col in REQUIRED_MINIMUM if col not in mapped_values]
            best = HeaderDetection(row_idx, mapped, raw_headers, missing, unknown)
            best_score = score
            best_exact = exact_count
    if best is None:
        raise ValueError('Não foi possível detectar a linha de cabeçalho da planilha.')
    if best_exact >= 12 and len(set(best.mapping.values()) & set(REQUIRED_MINIMUM)) >= 6:
        return best
    if best_score >= 15:
        return best
    raise ValueError('Não foi possível detectar a linha de cabeçalho da planilha.')

def detect_header(ws, max_scan_rows: int = 20):
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(getattr(ws, 'max_row', max_scan_rows), max_scan_rows), values_only=True):
        rows.append(list(row))
    return _tsiino_detect_header_from_values_v34(rows, max_scan_rows=max_scan_rows)

def is_strict_inpa_header_detection(detection) -> bool:
    vals = set(getattr(detection, 'mapping', {}).values())
    raws = getattr(detection, 'raw_headers', []) or []
    exact = sum(1 for raw in raws if _tsiino_map_header_v34(raw, allow_alias=False))
    return exact >= 12 and len(vals & set(REQUIRED_MINIMUM)) >= 6

# TSIINO_STANDARD_HEADER_DETECTOR_V36
# Detector robusto para a planilha padrão INPA/BRAHMS.
# Objetivo: uma linha com nomes técnicos reais (accession, collector, number, family, genus, ...)
# deve ser reconhecida como padrão mesmo se a primeira célula contiver texto auxiliar como
# "Nome do campo no Brahms ->". Cabeçalhos descritivos/português continuam indo para o mapeador.
def _tsiino_v36_norm(value):
    import re as _re
    import unicodedata as _unicodedata
    if value is None:
        return ''
    s = str(value).replace('\ufeff', '').strip()
    s = _unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not _unicodedata.combining(ch))
    s = s.lower()
    s = _re.sub(r'[^a-z0-9]+', '_', s).strip('_')
    return s

_TSIINO_V36_CANONICAL = list(CANONICAL_COLUMNS)
_TSIINO_V36_CANONICAL_SET = set(_TSIINO_V36_CANONICAL)
_TSIINO_V36_CANONICAL_BY_NORM = {_tsiino_v36_norm(c): c for c in _TSIINO_V36_CANONICAL}
_TSIINO_V36_REQUIRED_SET = set(REQUIRED_MINIMUM)
_TSIINO_V36_HELPER_HEADERS = {
    'nome_do_campo_no_brahms', 'nome_do_campo_no_brahms_', 'nome_do_campo_brahms',
    'recomendacao', 'recomendacoes', 'observacao', 'observacoes'
}


def _tsiino_v36_map_technical_header(raw, allow_alias=False):
    if raw is None:
        return None
    text = str(raw).replace('\ufeff', '').strip()
    if not text:
        return None
    if text in _TSIINO_V36_CANONICAL_SET:
        return text
    key = _tsiino_v36_norm(text)
    if key in _TSIINO_V36_CANONICAL_BY_NORM:
        return _TSIINO_V36_CANONICAL_BY_NORM[key]
    if allow_alias:
        try:
            return COLUMN_ALIASES.get(key)
        except Exception:
            return None
    return None


def _tsiino_v36_detect_header_from_values(rows, max_scan_rows=25):
    best = None
    best_score = -1
    best_exact = -1
    best_required_hits = -1

    for row_idx, row in enumerate((rows or [])[:max_scan_rows], start=1):
        raw_headers = []
        mapping = {}
        unknown = []
        exact_count = 0
        required_hits = set()
        canonical_hits = set()

        for cell in row:
            raw = '' if cell is None else str(cell).replace('\ufeff', '').strip()
            raw_headers.append(raw)
            if not raw:
                continue
            key = _tsiino_v36_norm(raw)
            exact = _tsiino_v36_map_technical_header(raw, allow_alias=False)
            # Para decidir se uma planilha e padrão, a evidência principal precisa ser exata.
            # Aliases só são usados como informação auxiliar para não quebrar compatibilidade interna.
            mapped = exact or _tsiino_v36_map_technical_header(raw, allow_alias=True)
            if exact:
                exact_count += 1
                canonical_hits.add(exact)
                if exact in _TSIINO_V36_REQUIRED_SET:
                    required_hits.add(exact)
            if mapped:
                mapping.setdefault(raw, mapped)
            elif key not in _TSIINO_V36_HELPER_HEADERS:
                unknown.append(raw)

        # Peso alto para nomes técnicos exatos. Isso impede que planilhas descritivas virem padrão.
        score = exact_count * 10 + len(required_hits) * 8 + len(canonical_hits) * 2
        if score > best_score:
            missing = [col for col in REQUIRED_MINIMUM if col not in set(mapping.values())]
            best = HeaderDetection(row_idx, mapping, raw_headers, missing, unknown)
            best_score = score
            best_exact = exact_count
            best_required_hits = len(required_hits)

    # Regra padrão: a planilha padrão INPA tem muitos campos técnicos exatos na mesma linha.
    # A planilha Zavatin anexada tem isso na linha 1: accession, collector, prefix, number, ...
    if best is not None and best_exact >= 8 and best_required_hits >= 5:
        return best

    # Fallback para templates muito reduzidos, mas ainda claramente técnicos.
    if best is not None and best_exact >= 5 and {'collector', 'number', 'family', 'genus'} <= set(best.mapping.values()):
        return best

    raise ValueError('Não foi possível detectar cabeçalho técnico INPA/BRAHMS. Abrir mapeador.')


def detect_header(ws, max_scan_rows: int = 25):
    rows = []
    max_row = min(getattr(ws, 'max_row', max_scan_rows) or max_scan_rows, max_scan_rows)
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        rows.append(list(row))
    return _tsiino_v36_detect_header_from_values(rows, max_scan_rows=max_scan_rows)


def is_strict_inpa_header_detection(detection) -> bool:
    vals = set(getattr(detection, 'mapping', {}).values() or [])
    raws = getattr(detection, 'raw_headers', []) or []
    exact = sum(1 for raw in raws if _tsiino_v36_map_technical_header(raw, allow_alias=False))
    return exact >= 8 and len(vals & set(REQUIRED_MINIMUM)) >= 5

